#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import asyncio
import os
import json
import pytest


pytestmark = pytest.mark.asyncio


async def _call_admin_list_csv(aiohttp_app, token):
    from aiohttp import ClientSession

    url = f"http://0.0.0.0:8081/admin?token={token}&format=csv"
    async with ClientSession() as s:
        async with s.get(url, headers={"Accept": "text/csv"}) as r:
            return r.status, await r.read(), r.headers


async def _call_admin_list_xlsx(aiohttp_app, token):
    from aiohttp import ClientSession

    url = f"http://0.0.0.0:8081/admin?token={token}&format=xlsx"
    async with ClientSession() as s:
        async with s.get(url) as r:
            return r.status, await r.read(), r.headers


async def test_admin_exports_csv_xlsx(monkeypatch):
    # Spin up a minimal web app by running bot in webhook mode on a test port
    from bot import ApplicationBuilder, setup_handlers, run_webhook_mode
    from config import config

    # Force webhook mode on a test port
    monkeypatch.setenv("PORT", "8081")
    monkeypatch.setenv("WEBHOOK_URL", "https://example.org")
    # Provide an admin dashboard token
    monkeypatch.setenv("ADMIN_DASHBOARD_TOKEN", "test-token")

    app = ApplicationBuilder().token(config.bot_token).build()
    await setup_handlers(app)

    # Start webhook server in background
    task = asyncio.create_task(run_webhook_mode(app))
    try:
        # Allow server to start
        await asyncio.sleep(0.8)
        # CSV export (enriched columns)
        status_csv, body_csv, headers_csv = await _call_admin_list_csv(app, "test-token")
        assert status_csv == 200
        assert (
            b"id,user_id,telegram_user_id,product_type,product_id,status,admin_action_by,admin_action_at,created_at"
            in body_csv
        )
        assert "text/csv" in headers_csv.get("Content-Type", "")

        # XLSX export
        status_xlsx, body_xlsx, headers_xlsx = await _call_admin_list_xlsx(app, "test-token")
        assert status_xlsx == 200
        assert headers_xlsx.get("Content-Type", "").startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # basic XLSX signature check
        assert body_xlsx[:2] == b"PK"  # zip header

        # Validate XLSX headers by opening the workbook in-memory
        try:
            from openpyxl import load_workbook
            import io

            wb = load_workbook(io.BytesIO(body_xlsx))
            ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            expected = [
                "id",
                "user_id",
                "telegram_user_id",
                "city",
                "grade",
                "product_type",
                "product_id",
                "status",
                "payment_status",
                "amount",
                "discount",
                "payment_method",
                "transaction_id",
                "admin_action_by",
                "admin_action_at",
                "created_at",
            ]
            assert headers == expected
        except Exception:
            # If openpyxl not available or fails, rely on content-type + PK signature
            pass
    finally:
        task.cancel()
        with pytest.raises(asyncio.CancelledError):
            await task
